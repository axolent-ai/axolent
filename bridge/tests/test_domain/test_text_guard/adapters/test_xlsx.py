"""Tests for the XLSX adapter (via TextGuardService).

These tests are skipped if openpyxl is not installed.
"""

from __future__ import annotations

from pathlib import Path

import pytest

try:
    from openpyxl import Workbook, load_workbook

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from application.text_guard_service import TextGuardService

pytestmark = pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")


@pytest.fixture
def tg_service() -> TextGuardService:
    """Text Guard service instance."""
    return TextGuardService()


@pytest.fixture
def sample_xlsx(tmp_path: Path) -> Path:
    """Create a sample XLSX with ASCII umlauts and a formula cell."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Das ist fuer dich."
    ws["A2"] = "Natuerlich moeglich."
    ws["A3"] = "=SUM(B1:B10)"  # Formula: must be skipped
    ws["A4"] = "This is correct English."
    filepath = tmp_path / "test.xlsx"
    wb.save(str(filepath))
    wb.close()
    return filepath


@pytest.fixture
def clean_xlsx(tmp_path: Path) -> Path:
    """Create an XLSX with correct umlauts."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Das ist für dich."
    filepath = tmp_path / "clean.xlsx"
    wb.save(str(filepath))
    wb.close()
    return filepath


class TestXlsxCheck:
    """Tests for checking XLSX files."""

    def test_finds_issues(
        self, tg_service: TextGuardService, sample_xlsx: Path
    ) -> None:
        """Detects ASCII umlauts in XLSX cells."""
        issues = tg_service.check_file(sample_xlsx, "de")
        assert len(issues) >= 2

    def test_skips_formulas(
        self, tg_service: TextGuardService, sample_xlsx: Path
    ) -> None:
        """Formula cells are not checked."""
        issues = tg_service.check_file(sample_xlsx, "de")
        formula_issues = [i for i in issues if "SUM" in i.excerpt]
        assert formula_issues == []

    def test_clean_file(self, tg_service: TextGuardService, clean_xlsx: Path) -> None:
        """Clean XLSX returns no issues."""
        issues = tg_service.check_file(clean_xlsx, "de")
        assert issues == []


class TestXlsxFix:
    """Tests for fixing XLSX files."""

    def test_fixes_in_place(
        self, tg_service: TextGuardService, sample_xlsx: Path
    ) -> None:
        """Fixes ASCII umlauts in XLSX in-place."""
        changed = tg_service.fix_file(sample_xlsx, "de")
        assert changed is True

        # Verify the fix
        wb = load_workbook(str(sample_xlsx))
        ws = wb.active
        assert "für" in ws["A1"].value
        assert ws["A3"].value == "=SUM(B1:B10)"  # Formula unchanged
        wb.close()

    def test_no_change_on_clean(
        self, tg_service: TextGuardService, clean_xlsx: Path
    ) -> None:
        """Clean XLSX is not modified."""
        changed = tg_service.fix_file(clean_xlsx, "de")
        assert changed is False
